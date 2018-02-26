import urllib.request, json 
import datetime
import time
import sys
from openpyxl import Workbook

# USAGE #
# py exchangeRates.py <start date> <end date> <base country> <countries to compare againgst>
#
# example: "py exchangeRates.py 2015-02-21 2015-11-30 USD GBP EUR"
#
# the above example would get the exchange rates between the U.S. dollar, the 
#    Great british Pound, and the Euro form Feb 21, 2015 till Nov 30, 2015

# DISCLAIMER #
# I hacked this together very quickly and didn't put any code in to check inputs. so just make
#    sure you enter everything correctly

# file to write to
outputFileName = "output.xlsx"

# base url stuff. 
URL_BASE = "https://api.fixer.io/"
BASE_COUNTRY_QUERY = "base="
SYMBOLS_QUERY_QUERY = "symbols="

# getting command line stuff
startDate = datetime.datetime.strptime(sys.argv[1], '%Y-%m-%d')
endDate = datetime.datetime.strptime(sys.argv[2], '%Y-%m-%d')
baseCountry = sys.argv[3]
countries = sys.argv[4:100]

# creating the workbook which will be the xlsx
# more info: https://openpyxl.readthedocs.io/en/stable/
wb = Workbook()


def formatDateString( year, month, day):
    return str(year) + "-" + str(month).zfill(2) + "-" + str(day).zfill(2)

def writeXlsxHeaders():
    ws = wb.active
    ws['A1'] = "DATE"
    for i in range(0, len(countries)):
        cell = chr(ord('A') + i + 1) + '1'
        ws[cell] = baseCountry + "->" + countries[i]


def writeToXlsx(date, rates, row):
    ws = wb.active
    ws['A' + str(row)] = date
    for i in range(0, len(rates)):
        cell = chr(ord('A') + i + 1) + str(row)
        ws[cell] = rates[countries[i]]
    return

writeXlsxHeaders()

date = startDate
row = 2
while date <= endDate:
    dateString = formatDateString(date.year, date.month, date.day)
    constructedUrl = URL_BASE + dateString + "?" + BASE_COUNTRY_QUERY + baseCountry + "&" + SYMBOLS_QUERY_QUERY + ",".join(countries)
    with urllib.request.urlopen(constructedUrl) as url:
        time.sleep(.25)
        print(dateString)
        data = json.loads(url.read())
        writeToXlsx(date, data['rates'], row)
    date += datetime.timedelta(days=1)
    row += 1

wb.save(outputFileName)